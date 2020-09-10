import collections
import os
import re
import traceback
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import historical_data_collection.excel_helpers as excel
from pprint import pprint
import config
import historical_data_collection.financial_statements_scraper.html_scraper_sec_edgar as html_scraper
# import historical_data_collection.financial_statements_scraper.xbrl_scraper_sec_edgar as xbrl_scraper


from zope.interface import Interface


class FinancialStatementsParserInterface(Interface):
    regex_patterns: dict

    def load_data_source(self, ticker: str) -> dict:
        """Load in the file links"""
        pass

    def scrape_tables(self, url: str, filing_date: datetime, filing_type: str) -> dict:
        """Extract tables from the currently loaded file."""
        pass

    def normalize_tables(self, regex_patterns, filing_date, input_dict, visited_data_names) -> (dict, dict):
        """Standardize tables to match across years and companies"""
        pass


date_regex = re.compile(r'^(0[1-9]|1[012])[- /.](0[1-9]|[12][0-9]|3[01])[- /.](19|20)\d\d$'  # match mm/dd/yyyy
                        r'|'
                        r'^(0[1-9]|[12][0-9]|3[01])[- /.](0[1-9]|1[012])[- /.](19|20)\d\d$'  # match dd-mm-yyyy
                        r'|'
                        r'^([^\s]+) (\d{2}),? ?(\d{4})$'  # match Month D, Yr (i.e. February 17, 2009 or February 17,2009)
                        r'|'
                        r'^\d{4}$'  # match year (i.e. 2011)
                        r'|'
                        'Fiscal\d{4}'
                        r'|'
                        r'^Change$'
                        r'|'
                        r'(\b\d{1,2}\D{0,3})?\b(?:Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|(Nov|Dec)(?:ember)?)\D?(\d{1,2}\D?)?\D?((19[7-9]\d|20\d{2})|\d{2})')


# make sure delta between bigger and smaller is 3 months i.e. 3 and 6, 6 and 9, 9 and 12
def quarterlize_statements(smaller_period_df, bigger_period_df, existing_quarters):
    global closest_smaller_period_index
    df_output = pd.DataFrame()
    for bigger_period_index, bigger_period_date in enumerate(bigger_period_df.columns):
        if bigger_period_date not in existing_quarters:
            try:
                closest_smaller_period_index = next(
                    index for (index, item) in enumerate(smaller_period_df.columns) if item < bigger_period_date)
            except:
                traceback.print_exc()
            # print(closest_smaller_period_index)
            df_output[bigger_period_date] = bigger_period_df[bigger_period_date] \
                                            - smaller_period_df.iloc[:, closest_smaller_period_index]
    # print(smaller_period_df.to_string())
    return df_output


def scrape_financial_statements(scraper_interface_implementation, ticker: str, how_many_years: int = 2,
                                how_many_quarters: int = 8):
    # if not FinancialStatementsParserInterface.implementedBy(scraper_interface_implementation):
    #     raise Exception

    global quarterly_df
    path = '{}/{}.xlsx'.format(config.FINANCIAL_STATEMENTS_DIR_PATH, ticker)
    log_folder_path = os.path.join(config.DATA_DIR_PATH, 'logs')
    if not os.path.exists(log_folder_path):
        os.mkdir(log_folder_path)
    company_log_path = os.path.join(log_folder_path, ticker)
    if not os.path.exists(company_log_path):
        os.mkdir(company_log_path)

    dictio_period_year_table = {}
    filing_dictio = scraper_interface_implementation().load_data_source(ticker=ticker)

    for filing_type, dates_and_links in filing_dictio.items():
        missing_dates_links = []
        existing_dates = excel.read_dates_from_csv(path, config.balance_sheet_yearly
        if filing_type == 'Yearly' else config.balance_sheet_quarterly)
        for date, link in dates_and_links:
            formatted_date = datetime.strptime(date, '%Y-%m-%d')
            if formatted_date not in existing_dates and formatted_date not in [x for x, y in missing_dates_links]:
                missing_dates_links.append((formatted_date, link))
        missing_dates_links.sort(key=lambda tup: tup[0], reverse=True)

        for index, (filing_date, link) in enumerate(missing_dates_links):
            try:
                if (index > how_many_years - 1 and filing_type == 'Yearly') \
                        or (index > how_many_quarters - 1 and filing_type == 'Quarterly'):
                    break
                print(filing_date, link)

                output = scraper_interface_implementation().scrape_tables(url=link, filing_date=filing_date,
                                                                          filing_type=filing_type)

                pprint(output)
                for sheet_period, sheet_dict in output.items():
                    if sheet_period not in dictio_period_year_table.keys():
                        dictio_period_year_table[sheet_period] = {}
                    for year, title_dict in sheet_dict.items():
                        # if we don't have the year in our dictio that collects everything, just add all and go to next year of the output
                        if year not in dictio_period_year_table[sheet_period].keys():
                            dictio_period_year_table[sheet_period][year] = title_dict
                            continue

                        #  else, we have a year, so we add up those two dicts together
                        for title, last_layer in title_dict.items():  # title / key:float
                            if title not in dictio_period_year_table[sheet_period][year].keys():
                                dictio_period_year_table[sheet_period][year][title] = last_layer

                            else:
                                dictio_period_year_table[sheet_period][year][title].update(last_layer)

            except Exception:
                traceback.print_exc()
        # pprint(dictio)

    import io
    import json
    with io.open(os.path.join(company_log_path, 'scraped_dictio.txt'), "w", encoding="utf-8") as f:
        f.write(json.dumps(str(dictio_period_year_table)))

    financials_dictio = {}

    for sheet_period, sheet_dict in dictio_period_year_table.items():
        visited_data_names = {}
        if sheet_period not in financials_dictio.keys():
            financials_dictio[sheet_period] = {}
        for year, title_dict in sheet_dict.items():
            if year not in financials_dictio[sheet_period].keys():
                financials_dictio[sheet_period][year] = {}
            visited_data_names, financials_dictio[sheet_period][year] = \
                scraper_interface_implementation().normalize_tables(
                    regex_patterns=scraper_interface_implementation().regex_patterns, filing_date=year,
                    input_dict=title_dict, visited_data_names=visited_data_names)

        # log = open(os.path.join(company_log_path, '{}_normalized_dictio.txt'.format(sheet_period)), "w")
        # print(visited_data_names, file=log)

    sheet_names = [config.balance_sheet_name, config.income_statement_name, config.cash_flow_statement_name]
    dfs_dictio = {}
    for sheet_name in sheet_names:
        if sheet_name not in dfs_dictio.keys():
            dfs_dictio[sheet_name] = {}
        first_balance_sheet_df = pd.DataFrame()
        visited_first_balance_sheet = False
        for sheet_period, sheet_dict in financials_dictio.items():
            if sheet_period not in dfs_dictio.keys():
                dfs_dictio[sheet_name][sheet_period] = pd.DataFrame()
                #  collections.OrderedDict(
            diction = collections.OrderedDict({i.strftime('%Y-%m-%d'): collections.OrderedDict({
                (j.split('_')[1], j.split('_')[-1])
                if 'Balance Sheet' not in sheet_name else (
                    (j.split('_')[1],
                     j.split('_')[2] if (len(j.split('_')) > 2) else ' ',
                     j.split('_')[-1]))
                : sheet_dict[i][j]
                for j in sheet_dict[i].keys() if j.split('_')[0] in sheet_name  # sheet name
            }) for i in sheet_dict.keys()})  # )  # date

            if sheet_name == config.balance_sheet_name and sheet_period in [config.yearly, config.quarterly] \
                    and not visited_first_balance_sheet:
                first_balance_sheet_df = pd.DataFrame.from_dict(diction)
                # print(first_balance_sheet_df.to_string())
                visited_first_balance_sheet = True
            else:
                df = pd.DataFrame.from_dict(diction)
                # pprint(diction)
                if sheet_name == config.balance_sheet_name and sheet_period in [config.yearly, config.quarterly]:
                    if visited_first_balance_sheet:
                        # print(df.to_string())
                        # df = pd.concat([first_balance_sheet_df, df], axis=0)
                        # first_balance_sheet_df.fillna(df, inplace=True)
                        # df.fillna(first_balance_sheet_df, inplace=True)
                        df = pd.concat([df, first_balance_sheet_df], axis=1, join='outer')
                        # print(df.to_string())
                        df = df.groupby(lambda x: x, axis=1).max()
                        year_dates = set([tuple_1 for tuple_1, tuple_2 in filing_dictio['Yearly']])
                        df_yearly = df[[x for x in df.columns if x in year_dates]]
                        for datafrme, period in zip([df_yearly, df], [config.yearly, config.quarterly]):
                            datafrme = datafrme.reindex(sorted(datafrme.columns, reverse=True), axis=1)
                            datafrme.dropna(axis=0, how='all', inplace=True)
                            datafrme = datafrme.loc[:, datafrme.any()]
                            if not datafrme.empty:
                                # dfs_dictio[sheet_name][sheet_period] = datafrme
                                excel.save_into_csv(path, datafrme, '{} ({})'.format(sheet_name, period))
                    break

                df = df.reindex(sorted(df.columns, reverse=True), axis=1)
                df.dropna(axis=0, how='all', inplace=True)
                df = df.loc[:, df.any()]
                if not df.empty:
                    # dfs_dictio[sheet_name][sheet_period] = df
                    excel.save_into_csv(path, df, '{} ({})'.format(sheet_name, sheet_period))

    #  this is to standardize cumulated 3 6 9 12 months
    for quarterly_statement in [config.cash_flow_statement_quarterly, config.income_statement_quarterly]:

        try:
            quarterly_df = pd.read_excel(path, quarterly_statement, index_col=[0, 1])
            # quarterly_df = dfs_dictio[]
        except:
            pass
        temp_statements = config.income_statements if quarterly_statement == config.income_statement_quarterly else config.cash_flow_statements
        for i, item in enumerate(temp_statements):
            try:
                smaller_period_df = pd.read_excel(path, item, index_col=[0, 1])
                bigger_period_df = pd.read_excel(path, temp_statements[i + 1], index_col=[0, 1])
                quarterly_df = pd.concat([quarterly_df,
                                          quarterlize_statements(smaller_period_df, bigger_period_df,
                                                                 quarterly_df.columns)],
                                         axis=1)
            except:
                pass
        try:
            quarterly_df = quarterly_df.reindex(sorted(quarterly_df.columns, reverse=True), axis=1)
            print(quarterly_df.to_string())
            excel.save_into_csv(path, quarterly_df, quarterly_statement, overwrite_sheet=True)
        except:
            pass

    #  then you can remove the 6 and 9 months sheets
    book = load_workbook(path)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        for sheet_name in writer.book.sheetnames:
            if config.six_months in sheet_name or config.nine_months in sheet_name:
                writer.book.remove(writer.book[sheet_name])
        writer.book.save(path)
        writer.close()

    #  finally, paint alternate rows to excel file
    book = load_workbook(path)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        if os.path.exists(path):
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    for sheet_name in writer.book.sheetnames:
        sheet = writer.book[sheet_name]
        df = pd.read_excel(path, sheet_name)

        start_column_space_for_index = 3 if config.balance_sheet_name in sheet_name else 2
        end_column_space_for_index = len(df.columns) + start_column_space_for_index - 1 \
            if config.balance_sheet_name not in sheet_name else len(df.columns) + start_column_space_for_index - 2
        row_space_for_index = 2
        for i in range(start_column_space_for_index, end_column_space_for_index):  # x-axis
            for j in range(row_space_for_index, len(df) + row_space_for_index):  # y-axis
                color = 'EEEEEE' if j % 2 == 0 else 'FFFFFF'  # integer odd-even alternation
                sheet.cell(row=j, column=i).fill = PatternFill(start_color=color,
                                                               fill_type='solid')

    writer.book.save(path)
    writer.close()


if __name__ == '__main__':
    tickers = excel.get_stock_universe('DJIA')
    for ticker in ['AAPL', 'FB']:
        scrape_financial_statements(scraper_interface_implementation=html_scraper.HtmlParser,
                                    ticker=ticker, how_many_years=3, how_many_quarters=0)
