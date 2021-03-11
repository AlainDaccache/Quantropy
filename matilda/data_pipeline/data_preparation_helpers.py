import collections
import os
import traceback
from datetime import datetime, timedelta
import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
import numpy as np
import xlrd

from matilda import config


def get_date_index(date, dates_values, lookback_index=0):
    if isinstance(dates_values[0], str):
        dates_values = [datetime.strptime(x, '%Y-%m-%d') for x in dates_values]
    elif isinstance(dates_values[0], np.datetime64):
        dates_values = [x.astype('M8[ms]').astype('O') for x in dates_values]

    if len(dates_values) > 1:
        if dates_values[0] > dates_values[1]:  # if dates decreasing rightwards or downwards
            date_index = next((index for (index, item) in enumerate(dates_values) if item < date), 0)
            # adjusted_lookback = date_item - lookback_period
            # lookback_index = next((
            #     index for (index, item) in enumerate(dates_values[date_index:]) if item <= adjusted_lookback), 0)
            return date_index + lookback_index
        else:  # if dates increasing rightwards or downwards
            date_index = next((index for (index, item) in enumerate(dates_values) if item > date), -1)
            # adjusted_lookback = date_item - lookback_period
            # lookback_index = next((
            #     index for (index, item) in enumerate(dates_values[date_index:]) if item > adjusted_lookback), -1)
            return date_index - lookback_index  # TODO Fix lookback index is a date here, convert before calling method

    else:
        return 0


def slice_series_dates(series, from_date, to_date):
    date_idx_from = get_date_index(from_date, series.index)
    date_idx_to = get_date_index(to_date, series.index)
    return series[date_idx_from:date_idx_to]


def save_into_csv(filename, df, sheet_name='Sheet1', startrow=None, overwrite_sheet=False, concat=False,
                  **to_excel_kwargs):
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # TODO Not working yet
        if concat and sheet_name in writer.book.sheetnames:
            try:
                sheet_df = pd.read_excel(filename, sheet_name,
                                         index_col=[0, 1, 2] if config.balance_sheet_name in sheet_name else [0, 1])
                print(sheet_df.to_string())
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])
                writer.book.create_sheet(sheet_name, idx)
                df = pd.concat([df, sheet_df], axis=1)
                df = df.reindex(sorted(df.columns, reverse=True), axis=1)
            except Exception:
                traceback.print_exc()

        # truncate sheet
        if overwrite_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def read_df_from_csv(path, sheet_name='Sheet1'):
    if os.path.exists(path):
        workbook = xlrd.open_workbook(path, on_demand=True)
        sheets = workbook.sheet_names()
        if sheet_name not in sheets:
            return pd.DataFrame()
        else:
            xls = pd.ExcelFile(path)
            return pd.read_excel(xls, sheet_name, index_col=0)
    return pd.DataFrame()


def read_entry_from_pickle(path, x, y, lookback_index=0):
    if os.path.exists(path):

        df: pd.DataFrame = pd.read_pickle(filepath_or_buffer=path)

        if isinstance(y, datetime):  # if the input is a date...
            date_index = get_date_index(date=y, dates_values=df.index.values, lookback_index=lookback_index)
            return df[x].iloc[date_index]

        elif isinstance(x, datetime):
            date_index = get_date_index(date=x, dates_values=df.columns, lookback_index=lookback_index)

            reduced_df = df.iloc[:, date_index]
            for el in list(y):
                if el in reduced_df.index:
                    reduced_df = reduced_df.loc[el]
                else:
                    return np.nan
            return reduced_df

        elif isinstance(y, list) and isinstance(y[0], datetime):
            to_return = pd.Series()
            for date in y:
                date_index = get_date_index(date=date, dates_values=df.index, lookback_index=lookback_index)

                reduced_df = df.iloc[date_index, :]
                for el in ([x] if not isinstance(x, list) else x):
                    if el in reduced_df.index:
                        reduced_df = reduced_df.loc[el]
                    else:
                        to_return[date] = np.nan
                to_return[date] = reduced_df
            return to_return

        elif isinstance(x, list) and isinstance(x[0], datetime):
            to_return = pd.Series()
            for date in x:
                date_index = get_date_index(date=date, dates_values=df.columns, lookback_index=lookback_index)
                reduced_df = df.iloc[:, date_index]
                if reduced_df.index.isin([tuple(y)]).any():
                    reduced_df = reduced_df.loc[tuple(y)]
                    to_return[date] = reduced_df
                else:
                    to_return[date] = np.nan

            return to_return

        else:
            return df[x].loc[y]
    else:
        return np.nan


def read_entry_from_csv(path, x, y, sheet_name='Sheet1', lookback_index=0, skip_first_sheet=False):
    if os.path.exists(path):
        # ticker = Path(path).stem
        if config.balance_sheet_name in sheet_name:
            index_col = [0, 1, 2]
        elif config.income_statement_name in sheet_name or config.cash_flow_statement_name in sheet_name:
            index_col = [0, 1]
        else:
            index_col = [0]

        df = pd.read_excel(pd.ExcelFile(path), sheet_name, index_col=index_col)

        if isinstance(y, datetime):  # if the input is a date...
            # if isinstance(df.index, pd.DatetimeIndex):
            date_index = get_date_index(date=y, dates_values=df.index.values, lookback_index=lookback_index)
            # print('The {} for {} on {}, lookback {}, is {}'.format(x, ticker, y, lookback_index, df[x].iloc[date_index]))
            return df[x].iloc[date_index]

        elif isinstance(x, datetime):
            date_index = get_date_index(date=x, dates_values=df.columns, lookback_index=lookback_index)

            reduced_df = df.iloc[:, date_index]
            for el in list(y):
                if el in reduced_df.index:
                    reduced_df = reduced_df.loc[el]
                else:
                    # print('The {} for {} on {}, lookback {}, is {}'.format(y, ticker, x, lookback_index, np.nan))
                    return np.nan
            # print('The {} for {} on {}, lookback {}, is {}'.format(y, ticker, x, lookback_index, reduced_df))
            return reduced_df
        else:
            # print('The {}/{} for {} is {}'.format(x, y, ticker, df[x].loc[y]))
            return df[x].loc[y]
    else:
        # print('The entry is {}'.format(np.nan))
        return np.nan


def read_dates_from_csv(path, sheet_name):
    if os.path.exists(path):
        sheets = xlrd.open_workbook(path, on_demand=True).sheet_names()
        if sheet_name not in sheets:
            return []
        with open(path, "r") as csv:
            xls = pd.ExcelFile(path)
            df = pd.read_excel(xls, '{}'.format(sheet_name), index_col=0)
            ls = []
            for col in df.columns:
                try:
                    ls.append(datetime.strptime(col, '%Y-%m-%d'))
                except:
                    continue
            return ls
    else:
        return []


def get_stock_universe(index='in_directory', date=datetime.now()):
    '''

    :param index: NASDAQ, DJIA, S&P1500, S&P500, S&P100, RUSSELL3000, RUSSELL2000, FTSE100
    :return:
    '''

    if index == 'in_directory':
        return [os.path.splitext(file)[0]
                for root, dirs, files in os.walk(config.FINANCIAL_STATEMENTS_DIR_PATH)
                for file in files]

    index_file_name = os.path.join(config.MARKET_TICKERS_DIR_PATH, 'Dow-Jones-Stock-Tickers.xlsx' if index == 'DJIA'
    else 'Russell-3000-Stock-Tickers.xlsx' if index == 'RUSSELL3000'
    else Exception)

    excel_df = pd.read_excel(index_file_name, index_col=0)
    date_index = get_date_index(date=date, dates_values=excel_df.index)
    tickers = excel_df.iloc[date_index]
    return tickers.to_list()


def slice_resample_merge_returns(returns: list, from_date=None, to_date=None, lookback=None,
                                 frequency: str = 'Monthly'):
    '''

    :param returns:
    :param from_date:
    :param to_date:
    :param lookback:
    :return:
    '''

    # Load each asset returns and merge
    returns_copy = []
    for retrn in returns:
        if isinstance(retrn, str):
            path = '{}/{}.xlsx'.format(config.STOCK_PRICES_DIR_PATH, retrn)
            series = read_df_from_csv(path)['Adj Close'].pct_change().rename(retrn)
            returns_copy.append(series)
        elif isinstance(retrn, pd.Series):
            returns_copy.append(retrn)
        elif isinstance(retrn, pd.DataFrame):
            for col in retrn.columns:
                returns_copy.append(retrn[col])
        else:
            raise Exception

    # Resample based on desired frequency and merge
    merged_returns = pd.DataFrame()
    for retrn in returns_copy:
        resampled_returns = retrn.resample(frequency[0]).apply(lambda x: ((x + 1).cumprod() - 1).last("D"))
        # Resample usually resets date to beginning of day, so we re-do the end of day trick:
        resampled_returns.index = resampled_returns.index + timedelta(days=1) - timedelta(seconds=1)
        merged_returns = merged_returns.join(resampled_returns.to_frame(), how='outer')

    # Cutoff based on from_date and to_date

    # Go over returns list because merged would have nans
    to_date = min([series.index[-1] for series in returns_copy]) if to_date is None else to_date
    to_date_idx = get_date_index(date=to_date, dates_values=merged_returns.index)

    if from_date is not None:  # from_date has precedence over lookback if both are not none
        from_date_idx = get_date_index(date=from_date, dates_values=merged_returns.index)
    elif lookback is not None:
        if isinstance(lookback, int) and frequency is not None:
            period_to_int = {'D': 1, 'W': 7, 'M': 30.5, 'Y': 365.25}
            lookback = timedelta(days=int(period_to_int[frequency] * lookback))
        elif not isinstance(lookback, timedelta):
            raise Exception
        from_date_idx = get_date_index(date=to_date - lookback, dates_values=merged_returns.index)
    else:
        from_date_idx = 0

    merged_returns = merged_returns.iloc[from_date_idx:to_date_idx]

    for col in merged_returns.columns:
        merged_returns[col] = merged_returns[col].apply(lambda y: 0 if isinstance(y, np.ndarray) else y)

    return merged_returns


def fill_last_level(dictionary, fill_value=0):
    if isinstance(dictionary, dict):
        result = []
        for v in dictionary.values():
            result.extend(fill_last_level(v))
        return result
    else:
        return [dictionary]


def unflatten(dictionary):
    resultDict = dict()
    for key, value in dictionary.items():
        parts = key.split("_")
        d = resultDict
        for part in parts[:-1]:
            try:
                if part not in d:
                    d[part] = dict()
                d = d[part]
            except:
                continue
        try:
            d[parts[-1]] = value
        except:
            continue
    return resultDict


def flatten_dict(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = parent_key + sep + k if parent_key else k
        if isinstance(v, collections.abc.MutableMapping):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


def save_pretty_excel(path, financials_dictio, with_pickle=True):
    for sheet_name in [config.income_statement_name, config.cash_flow_statement_name]:
        for sheet_period, sheet_dict in financials_dictio.items():

            # keep 2 levels for the Income Statement and Cash Flow i.e.
            #   Operating Expenses -> Research and Development Expense -> Value
            #   Cash Flow from Financing Activities -> Dividend Payments -> Value

            diction = collections.OrderedDict({i: collections.OrderedDict({
                (j.split('_')[1], j.split('_')[-1]): sheet_dict[i][j]
                for j in sheet_dict[i].keys() if j.split('_')[0] in sheet_name  # sheet name
            }) for i in sheet_dict.keys()})  # )  # date

            df = pd.DataFrame.from_dict(diction)
            df = df.reindex(sorted(df.columns, reverse=True), axis=1)
            df.dropna(axis=0, how='all', inplace=True)  # either NA
            df = df[(df.T != 0).any()]  # or 0
            df = df.loc[:, df.any()]
            if not df.empty:
                if with_pickle:
                    stock = path.split('/')[-1].split('.xlsx')[0]
                    pickly_path = os.path.join(config.FINANCIAL_STATEMENTS_DIR_PATH_PICKLE, sheet_period,
                                               sheet_name, '{}.pkl'.format(stock))
                    df.to_pickle(pickly_path)
                save_into_csv(path, df, '{} ({})'.format(sheet_name, sheet_period))

    #  this is to standardize cumulated 3 6 9 12 months
    # for quarterly_statement in [config.cash_flow_statement_quarterly, config.income_statement_quarterly]:
    #
    #     try:
    #         quarterly_df = pd.read_excel(path, quarterly_statement, index_col=[0, 1])
    #     except:
    #         pass
    #     temp_statements = config.income_statements if quarterly_statement == config.income_statement_quarterly else config.cash_flow_statements
    #     for i, item in enumerate(temp_statements):
    #         try:
    #             smaller_period_df = pd.read_excel(path, item, index_col=[0, 1])
    #             bigger_period_df = pd.read_excel(path, temp_statements[i + 1], index_col=[0, 1])
    #             quarterly_df = pd.concat([quarterly_df,
    #                                       quarterlize_statements(smaller_period_df, bigger_period_df,
    #                                                              quarterly_df.columns)],
    #                                      axis=1)
    #         except:
    #             pass
    # try:
    #     quarterly_df = quarterly_df.reindex(sorted(quarterly_df.columns, reverse=True), axis=1)
    #     print(quarterly_df.to_string())
    #     save_into_csv(path, quarterly_df, quarterly_statement, overwrite_sheet=True)
    # except:
    #     pass

    #  then you can remove the 6 and 9 months sheets
    book = load_workbook(path)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        for sheet_name in writer.book.sheetnames:
            if config.six_months in sheet_name or config.nine_months in sheet_name:
                writer.book.remove(writer.book[sheet_name])
        writer.book.save(path)
        writer.close()

    # special case for balance sheet (no cumulative so can skip the whole 6/9 months standardization)
    balance_sheet_df = pd.DataFrame()
    for sheet_name in [config.balance_sheet_name]:
        for sheet_period in [config.yearly, config.quarterly]:
            # only keep 3 levels of dictionary for Balance Sheet i.e.
            #   Assets -> Current Assets -> Cash Equivalents -> Value
            #   Liabilities & Shareholders' Equity -> Liabilities -> Accounts Payable -> Value

            diction = collections.OrderedDict(
                {i: collections.OrderedDict({(entry_name.split('_')[1],
                                              entry_name.split('_')[2] if (len(entry_name.split('_')) > 2) else ' ',
                                              entry_name.split('_')[-1]): financials_dictio[sheet_period][i][entry_name]
                                             for entry_name in financials_dictio[sheet_period][i].keys() if
                                             entry_name.split('_')[0] in sheet_name
                                             }) for i in financials_dictio[sheet_period].keys()})  # date

            # Because Balance Sheet is a statement of position, the Quarterly statement that has the same
            # date as the Yearly statement, will have the same information. Sometimes reports avoid explicitly
            # specifying a balance sheet for a certain quarter for that reason, so we need to consider the case.
            balance_sheet_df = pd.concat([balance_sheet_df, pd.DataFrame.from_dict(diction)], axis=1, join='outer')

        # since we might have duplicate columns after concatenating, we drop the duplicated ones
        balance_sheet_df = balance_sheet_df.loc[:, ~balance_sheet_df.columns.duplicated()]
        # keep the Yearly balance sheets as a separate df
        balance_sheet_yearly = balance_sheet_df[financials_dictio['Yearly'].keys()]  # now can extract the Yearly

        for datafrme, period in zip([balance_sheet_yearly, balance_sheet_df], [config.yearly, config.quarterly]):
            datafrme = datafrme.reindex(sorted(datafrme.columns, reverse=True), axis=1)
            datafrme.dropna(axis=0, how='all', inplace=True)
            datafrme = datafrme.loc[:, datafrme.any()]
            if not datafrme.empty:
                if with_pickle:
                    stock = path.split('/')[-1].split('.xlsx')[0]
                    pickly_path = os.path.join(config.FINANCIAL_STATEMENTS_DIR_PATH_PICKLE, period,
                                               sheet_name, '{}.pkl'.format(stock))
                    datafrme.to_pickle(pickly_path)
                save_into_csv(path, datafrme, '{} ({})'.format(sheet_name, period))

    #  finally, paint alternate rows to excel file
    book = load_workbook(path)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        if os.path.exists(path):
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    for sheet_name in writer.book.sheetnames:
        sheet = writer.book[sheet_name]
        df = pd.read_excel(path, sheet_name)

        start_column_space_for_index = 3 if config.balance_sheet_name in sheet_name else 2
        end_column_space_for_index = len(df.columns) + start_column_space_for_index - 1 \
            if config.balance_sheet_name not in sheet_name else len(df.columns) + start_column_space_for_index - 2
        row_space_for_index = 2
        for i in range(start_column_space_for_index, end_column_space_for_index):  # x-axis
            for j in range(row_space_for_index, len(df) + row_space_for_index):  # y-axis
                color = 'EEEEEE' if j % 2 == 0 else 'FFFFFF'  # integer odd-even alternation
                sheet.cell(row=j, column=i).fill = PatternFill(start_color=color,
                                                               fill_type='solid')

    writer.book.save(path)
    writer.close()


if __name__ == '__main__':
    print(get_stock_universe('DJIA'))
