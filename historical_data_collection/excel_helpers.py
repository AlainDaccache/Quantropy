import collections
import os
import traceback
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
import config
from openpyxl import load_workbook
import numpy as np
import xlrd
from typing import Callable


def get_date_index(date, dates_values, lookback_index=0):
    if isinstance(dates_values[0], str):
        dates_values = [datetime.strptime(x, '%Y-%m-%d') for x in dates_values]
    elif isinstance(dates_values[0], np.datetime64):
        dates_values = [x.astype('M8[ms]').astype('O') for x in dates_values]

    if len(dates_values) > 1:
        if dates_values[0] > dates_values[1]:  # if dates decreasing rightwards or downwards
            date_index = next((index for (index, item) in enumerate(dates_values) if item < date), 0)
            # adjusted_lookback = date_item - lookback_period
            # lookback_index = next((
            #     index for (index, item) in enumerate(dates_values[date_index:]) if item <= adjusted_lookback), 0)
            return date_index + lookback_index
        else:  # if dates increasing rightwards or downwards
            date_index = next((index for (index, item) in enumerate(dates_values) if item > date), -1)
            # adjusted_lookback = date_item - lookback_period
            # lookback_index = next((
            #     index for (index, item) in enumerate(dates_values[date_index:]) if item > adjusted_lookback), -1)
            return date_index - lookback_index  # TODO Fix lookback index is a date here, convert before calling method

    else:
        return 0


def slice_series_dates(series, from_date, to_date):
    date_idx_from = get_date_index(from_date, series.index)
    date_idx_to = get_date_index(to_date, series.index)
    return series[date_idx_from:date_idx_to]


def save_into_csv(filename, df, sheet_name='Sheet1', startrow=None,
                  overwrite_sheet=False, concat=False,
                  **to_excel_kwargs):
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # TODO Not working yet
        if concat and sheet_name in writer.book.sheetnames:
            try:
                sheet_df = pd.read_excel(filename, sheet_name,
                                         index_col=[0, 1, 2] if config.balance_sheet_name in sheet_name else [0, 1])
                print(sheet_df.to_string())
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])
                writer.book.create_sheet(sheet_name, idx)
                df = pd.concat([df, sheet_df], axis=1)
                df = df.reindex(sorted(df.columns, reverse=True), axis=1)
            except:
                traceback.print_exc()

        # truncate sheet
        if overwrite_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def read_df_from_csv(path, sheet_name='Sheet1'):
    if os.path.exists(path):
        workbook = xlrd.open_workbook(path, on_demand=True)
        sheets = workbook.sheet_names()
        if sheet_name not in sheets:
            return pd.DataFrame()
        else:
            xls = pd.ExcelFile(path)
            return pd.read_excel(xls, sheet_name, index_col=0)
    return pd.DataFrame()


def read_entry_from_csv(path, x, y, sheet_name='Sheet1', lookback_index=0, skip_first_sheet=False):
    if os.path.exists(path):
        ticker = Path(path).stem
        if config.balance_sheet_name in sheet_name:
            index_col = [0, 1, 2]
        elif config.income_statement_name in sheet_name or config.cash_flow_statement_name in sheet_name:
            index_col = [0, 1]
        else:
            index_col = [0]

        df = pd.read_excel(pd.ExcelFile(path), sheet_name, index_col=index_col)

        if isinstance(y, datetime):  # if the input is a date...
            # if isinstance(df.index, pd.DatetimeIndex):
            date_index = get_date_index(date=y, dates_values=df.index.values, lookback_index=lookback_index)
            # print('The {} for {} on {}, lookback {}, is {}'.format(x, ticker, y, lookback_index, df[x].iloc[date_index]))
            return df[x].iloc[date_index]

        elif isinstance(x, datetime):
            date_index = get_date_index(date=x, dates_values=df.columns, lookback_index=lookback_index)

            reduced_df = df.iloc[:, date_index]
            for el in list(y):
                if el in reduced_df.index:
                    reduced_df = reduced_df.loc[el]
                else:
                    # print('The {} for {} on {}, lookback {}, is {}'.format(y, ticker, x, lookback_index, np.nan))
                    return np.nan
            # print('The {} for {} on {}, lookback {}, is {}'.format(y, ticker, x, lookback_index, reduced_df))
            return reduced_df
        else:
            # print('The {}/{} for {} is {}'.format(x, y, ticker, df[x].loc[y]))
            return df[x].loc[y]
    else:
        # print('The entry is {}'.format(np.nan))
        return np.nan


def read_dates_from_csv(path, sheet_name):
    if os.path.exists(path):
        sheets = xlrd.open_workbook(path, on_demand=True).sheet_names()
        if sheet_name not in sheets:
            return []
        with open(path, "r") as csv:
            xls = pd.ExcelFile(path)
            df = pd.read_excel(xls, '{}'.format(sheet_name), index_col=0)
            ls = []
            for col in df.columns:
                try:
                    ls.append(datetime.strptime(col, '%Y-%m-%d'))
                except:
                    continue
            return ls
    else:
        return []


def get_stock_universe(index='in_directory', date=datetime.now()):
    '''

    :param index: NASDAQ, DJIA, S&P1500, S&P500, S&P100, RUSSELL3000, RUSSELL2000, FTSE100
    :return:
    '''

    if index == 'in_directory':
        return [os.path.splitext(file)[0]
                for root, dirs, files in os.walk(config.FINANCIAL_STATEMENTS_DIR_PATH)
                for file in files]

    index_file_name = os.path.join(config.MARKET_TICKERS_DIR_PATH, 'Dow-Jones-Stock-Tickers.xlsx' if index == 'DJIA'
    else 'Russell-3000-Stock-Tickers.xlsx' if index == 'RUSSELL3000'
    else Exception)

    excel_df = pd.read_excel(index_file_name, index_col=0)
    date_index = get_date_index(date=date, dates_values=excel_df.index)
    tickers = excel_df.iloc[date_index]
    return tickers.to_list()


def slice_resample_merge_returns(portfolio, benchmark=None,
                                 from_date=None, to_date=None,
                                 period='Monthly'):
    if isinstance(portfolio, str):
        path = '{}/{}.xlsx'.format(config.STOCK_PRICES_DIR_PATH, portfolio)
        portfolio_returns = read_df_from_csv(path)['Adj Close'].pct_change()
    elif isinstance(portfolio, pd.Series):
        portfolio_returns = portfolio
    else:
        raise Exception
    # Slice and resample portfolio (or asset) returns

    if to_date is None:
        to_date = portfolio_returns.index[-1]
    if from_date is None:
        from_date = to_date - timedelta(days=5 * 365)

    date_idx_from = get_date_index(from_date, portfolio_returns.index)
    date_idx_to = get_date_index(to_date, portfolio_returns.index)
    portfolio_returns = portfolio_returns[date_idx_from:date_idx_to]
    portfolio_returns = portfolio_returns.resample(period[0]).apply(lambda x: ((x + 1).cumprod() - 1).last("D"))
    portfolio_returns = portfolio_returns.apply(lambda y: 0 if isinstance(y, np.ndarray) else y)
    # Resample usually resets date to beginning of day, so we re-do the end of day trick:
    portfolio_returns.index = portfolio_returns.index + timedelta(days=1) - timedelta(seconds=1)

    if benchmark is None:
        df = read_df_from_csv(path='{}/{}.xlsx'.format(config.FACTORS_DIR_PATH, 'CAPM'), sheet_name=period)
        benchmark_returns = df['Mkt-RF'] + df['RF']

    elif isinstance(benchmark, str):
        df = read_df_from_csv(path='{}/{}.xlsx'.format(config.STOCK_PRICES_DIR_PATH, benchmark))
        benchmark_returns = df['Adj Close'].pct_change()
    elif isinstance(benchmark, pd.Series):
        benchmark_returns = benchmark
    else:
        raise Exception
    benchmark_returns = benchmark_returns.apply(lambda y: 0 if isinstance(y, np.ndarray) else y)
    benchmark_returns = benchmark_returns.resample(period[0]).apply(lambda x: ((x + 1).cumprod() - 1).last("D"))
    # Resample usually resets date to beginning of day, so we re-do the end of day trick:
    benchmark_returns.index = benchmark_returns.index + timedelta(days=1) - timedelta(seconds=1)

    benchmark_returns.rename('Benchmark', inplace=True)
    portfolio_returns.rename('Portfolio', inplace=True)
    merged_df = pd.concat([benchmark_returns, portfolio_returns], axis=1).dropna()
    benchmark_returns, portfolio_returns = merged_df['Benchmark'], merged_df['Portfolio']
    benchmark_returns = benchmark_returns.apply(lambda y: 0 if isinstance(y, np.ndarray) else y)
    portfolio_returns = portfolio_returns.apply(lambda y: 0 if isinstance(y, np.ndarray) else y)
    return benchmark_returns, portfolio_returns


def unflatten(dictionary):
    resultDict = dict()
    for key, value in dictionary.items():
        parts = key.split("_")
        d = resultDict
        for part in parts[:-1]:
            try:
                if part not in d:
                    d[part] = dict()
                d = d[part]
            except:
                continue
        try:
            d[parts[-1]] = value
        except:
            continue
    return resultDict


def flatten_dict(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = parent_key + sep + k if parent_key else k
        if isinstance(v, collections.abc.MutableMapping):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


if __name__ == '__main__':
    print(get_stock_universe('DJIA'))
