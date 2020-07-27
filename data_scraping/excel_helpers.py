import os
import traceback
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill, Font

import config
import data_scraping.financial_data_scraper as scraper
from openpyxl import load_workbook
import numpy as np
import xlrd
from typing import Callable


def get_date_index(date, dates_values, lookback_index=0):
    if isinstance(dates_values[0], str):
        dates_values = [datetime.strptime(x, '%Y-%m-%d') for x in dates_values]
    elif isinstance(dates_values[0], np.datetime64):
        dates_values = [x.astype('M8[ms]').astype('O') for x in dates_values]
    if len(dates_values) > 1:
        if dates_values[0] > dates_values[1]:  # if dates decreasing rightwards or downwards
            date_index, date_item = next(((index, item) for (index, item) in enumerate(dates_values) if item < date), 0)
            # adjusted_lookback = date_item - lookback_period
            # lookback_index = next((
            #     index for (index, item) in enumerate(dates_values[date_index:]) if item <= adjusted_lookback), 0)
            return date_index + lookback_index
        else:  # if dates increasing rightwards or downwards
            date_index, date_item = next(((index, item) for (index, item) in enumerate(dates_values) if item > date),
                                         -1)
            # adjusted_lookback = date_item - lookback_period
            # lookback_index = next((
            #     index for (index, item) in enumerate(dates_values[date_index:]) if item > adjusted_lookback), -1)
            return date_index - lookback_index
    else:
        return 0


def save_into_csv(filename, df, sheet_name='Sheet1', startrow=None,
                  overwrite_sheet=False, concat=False,
                  **to_excel_kwargs):

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # TODO Not working yet
        if concat and sheet_name in writer.book.sheetnames:
            try:
                sheet_df = pd.read_excel(filename, sheet_name, index_col=[0, 1, 2] if config.balance_sheet_name in sheet_name else [0, 1])
                print(sheet_df.to_string())
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])
                writer.book.create_sheet(sheet_name, idx)
                df = pd.concat([df, sheet_df], axis=1)
                df = df.reindex(sorted(df.columns, reverse=True), axis=1)
            except:
                traceback.print_exc()

        # truncate sheet
        if overwrite_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def read_df_from_csv(path, sheet_name):
    if os.path.exists(path):
        workbook = xlrd.open_workbook(path, on_demand=True)
        sheets = workbook.sheet_names()
        if sheet_name not in sheets:
            return pd.DataFrame()
        else:
            xls = pd.ExcelFile(path)
            return pd.read_excel(xls, sheet_name, index_col=0)
    return pd.DataFrame()


def read_entry_from_csv(path, sheet_name, x, y, lookback_index=0, skip_first_sheet=False):
    if os.path.exists(path):
        # maybe the file exists, but the sheet is not in the file
        stock = Path(path).stem
        sheets = xlrd.open_workbook(path, on_demand=True).sheet_names()
        if sheet_name not in sheets:
            if sheet_name == config.stock_prices_sheet_name:
                scraper.get_stock_prices(stock)
            elif sheet_name in config.quarterly_statements:
                # scraper.scrape_financial_statements(stock, '10-Q')
                return Exception
            elif sheet_name in config.yearly_statements:
                scraper.scrape_financial_statements(stock, '10-K')

    # if the file doesn't exist
    else:
        if config.FINANCIAL_STATEMENTS_DIR_PATH in path:
            stock = Path(path).stem
            scraper.get_stock_prices(stock)
            scraper.scrape_financial_statements(stock, '10-K')
            # scraper.scrape_financial_statements(stock, '10-Q')
        # elif path == config.beta_factors_file_path:
        #     scraper.get_beta_factors()

    xls = pd.ExcelFile(path)
    if config.balance_sheet_name in sheet_name:
        index_col = [0, 1, 2]
    elif config.income_statement_name in sheet_name or config.cash_flow_statement_name in sheet_name:
        index_col = [0, 1]
    else:
        index_col = [0]

    df = pd.read_excel(xls, sheet_name, index_col=index_col)
    nan_row = np.array(df.isnull().any(1)).nonzero()[0]
    if not skip_first_sheet:
        pass
    else:
        pass
    if isinstance(y, datetime):  # if the input is a date...
        # if isinstance(df.index, pd.DatetimeIndex):
        date_index = get_date_index(date=y, dates_values=df.index.values, lookback_index=lookback_index)
        return df[x].iloc[date_index]

    elif isinstance(x, datetime):
        date_index = get_date_index(date=x, dates_values=df.columns, lookback_index=lookback_index)

        reduced_df = df.iloc[:, date_index]
        for el in list(y):
            if el in reduced_df.index:
                reduced_df = reduced_df.loc[el]
            else:
                print('{} for {} not found!'.format(y, x))
                return np.nan
        return reduced_df
    else:
        return df[x].loc[y]


def read_dates_from_csv(path, sheet_name):
    if os.path.exists(path):
        sheets = xlrd.open_workbook(path, on_demand=True).sheet_names()
        if sheet_name not in sheets:
            return []
        with open(path, "r") as csv:
            xls = pd.ExcelFile(path)
            df = pd.read_excel(xls, '{}'.format(sheet_name), index_col=0)
            ls = []
            for col in df.columns:
                try:
                    ls.append(datetime.strptime(col, '%Y-%m-%d'))
                except:
                    continue
            return ls
    else:
        return []


def get_stock_universe():
    tickers = []
    directory = config.FINANCIAL_STATEMENTS_DIR_PATH
    for root, dirs, files in os.walk(directory):
        for file in files:
            tickers.append(os.path.splitext(file)[0])
    return tickers


def get_industry(stock: str):
    pass


def companies_in_industry(industry: str):
    df = pd.read_csv('companies_overview.csv', index_col='Ticker')
    ind = df['Industry'] == industry
    tickers_in_industry = [ticker for ticker, boolean in ind.iteritems() if boolean]
    return tickers_in_industry


def metric_in_industry(metric: Callable, industry: str) -> pd.Series(dtype="float64"):
    output = pd.Series(dtype='float64')
    tickers_in_industry = companies_in_industry(industry)
    for ticker in tickers_in_industry:
        output[ticker] = metric(ticker)
    return output


# get companies greater or lower than a certain percentile regarding metrics
# for Q1, percentile is 25, for Q2, percentile is 50, for Q3, percentile is 75
# comparator: '>', '<'
def companies_comparative_metrics(metrics_series: pd.Series(dtype='float64'), percentile: int, comparator: str):
    quantile = np.percentile(metrics_series, percentile)
    if comparator == '>':
        output = [ticker for ticker, metric_value in metrics_series.iteritems() if metric_value > quantile]
    else:
        output = [ticker for ticker, metric_value in metrics_series.iteritems() if metric_value < quantile]
    return output


if __name__ == '__main__':
    # metrics_series = metric_in_industry(partial(fi.earnings_per_share, ttm=False), 'SERVICES-COMPUTER PROGRAMMING, DATA PROCESSING, ETC.')
    # companies = companies_comparative_metrics(metrics_series, percentile=25, comparator='>')
    # print(companies)
    financial_path = '{}/{}.xlsx'.format(config.FINANCIAL_STATEMENTS_DIR_NAME, 'TSLA')
    # print(read_entry_from_csv(financial_path, config.stock_prices_sheet_name, datetime.now(), 'Adj Close'))
    # print(read_entry_from_csv(financial_path, config.balance_sheet_yearly, datetime(2018, 4, 4), 'Restricted Cash Non Current'))
    print(read_entry_from_csv(config.FACTORS_DIR_PATH, config.yearly_factors, datetime(2001, 1, 1), 'RF'))
