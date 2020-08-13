import collections
import csv
import math
import os
import pickle
import re
import sys
import traceback
import urllib.request
import zipfile
from datetime import datetime, timedelta
import typing
from time import sleep
import numpy as np
import pandas as pd
import requests
import unicodedata
from bs4 import BeautifulSoup, NavigableString
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import data_scraping.regex_patterns as fin_reg
import pandas_datareader.data as web
import data_scraping.excel_helpers as excel
from pprint import pprint
import tabula
import ta
import config
from titlecase import titlecase
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from typing import Dict, List, Union

'''
Beautiful BeautifulSoup Usage

html = urllib2.urlopen(url).read()
bs = BeautifulSoup(html)
table = bs.find(lambda tag: tag.name=='table' and tag.has_attr('id') and tag['id']=="Table1") 
rows = table.findAll(lambda tag: tag.name=='tr')
'''


# TODO: Sometimes only year format is available, current implementation skips
# TODO: What if both quarterly and yearly are in same table? first_level is wrong
# TODO: Save filings urls in txt files
# TODO: Check if normalization round three can take shares outstanding from income statements
# TODO: Make this into an interface
# TODO: Add ?!.*_ trick in regexes

# class ParserInterface:
#     def load_financial_statements_links(self, ticker: str) -> Dict[str, List[(str, str)]]:
#         pass
#
#     def parse_tables(self, link: str) -> Dict[str, Dict[str, float]]:
#         pass
#
#     def standardize_tables(self):
#         pass


def get_company_cik(ticker):
    URL = 'http://www.sec.gov/cgi-bin/browse-edgar?CIK={}&Find=Search&owner=exclude&action=getcompany'.format(ticker)
    response = requests.get(URL)
    CIK_RE = re.compile(r'.*CIK=(\d{10}).*')
    cik = CIK_RE.findall(response.text)[0]
    print('Company CIK for {} is {}'.format(ticker, cik))
    return cik


def get_filings_urls_first_layer(cik, filing_type):
    base_url = "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK={}&type={}".format(cik, filing_type)
    edgar_resp = requests.get(base_url).text
    print(base_url)
    soup = BeautifulSoup(edgar_resp, 'html.parser')
    table_tag = soup.find('table', class_='tableFile2')
    rows = table_tag.find_all('tr')
    doc_links = []
    for row in rows[1:]:
        cells = row.find_all('td')
        doc_links.append('https://www.sec.gov' + cells[1].a['href'])
    return doc_links


def get_filings_urls_second_layer(doc_links):
    links_dictionary = {
        'XBRL Links': [],  # for new submissions
        'HTML Links': [],  # if no XML, then we put HTML (for older submissions).
        'TXT Links': []  # if no HTML, then we put TXT (for even older submissions)
    }

    for doc_link in doc_links:
        # Obtain HTML for document page
        doc_resp = requests.get(doc_link).text
        # Find the XBRL link
        soup = BeautifulSoup(doc_resp, 'html.parser')
        # first, find period of report
        head_divs = soup.find_all('div', class_='infoHead')
        cell_index = next((index for (index, item) in enumerate(head_divs) if item.text == 'Period of Report'), -1)

        period_of_report = ''
        try:
            siblings = head_divs[cell_index].next_siblings
            for sib in siblings:
                if isinstance(sib, NavigableString):
                    continue
                else:
                    period_of_report = sib.text
                    break
        except:
            traceback.print_exc()

        # first, try finding a XML document
        table_tag = soup.find('table', class_='tableFile', summary='Data Files')
        no_xml_link = True
        if table_tag is not None:
            rows = table_tag.find_all('tr')
            for row_index, row in enumerate(rows[1:]):
                cells = row.find_all('td')
                link = 'https://www.sec.gov' + cells[2].a['href']
                if 'XML' in cells[3].text or 'INS' in cells[3].text:
                    links_dictionary['XBRL Links'].append((period_of_report, link))
                    no_xml_link = False

        # if no XML document found, try finding html documents
        no_html_link = True
        no_xml_link = True  # this is to include XML links as HTMLs to test HTML scraping, so remove when done with xml
        if no_xml_link:
            table_tag = soup.find('table', class_='tableFile', summary='Document Format Files')
            rows = table_tag.find_all('tr')
            for row in rows[1:]:
                cells = row.find_all('td')
                link = 'https://www.sec.gov' + cells[2].a['href']
                if 'htm' in cells[2].text and cells[3].text in ['10-K', '10-Q']:
                    links_dictionary['HTML Links'].append((period_of_report, link))
                    no_html_link = False

        # if no HTML document found, then get the txt
        if no_html_link and no_xml_link:
            table_tag = soup.find('table', class_='tableFile', summary='Document Format Files')
            rows = table_tag.find_all('tr')
            for row in rows[1:]:
                cells = row.find_all('td')
                link = 'https://www.sec.gov' + cells[2].a['href']
                if cells[1].text == 'Complete submission text file':
                    links_dictionary['TXT Links'].append((period_of_report, link))

    return links_dictionary


def is_bold(row, alltext=False):
    soup = BeautifulSoup(str(row), features='html.parser')
    bolded_row_text = soup.find_all(
        lambda tag: tag.name == 'b' or (
                tag.has_attr('style') and (re.search('bold|font-weight:700', str(tag['style']), re.IGNORECASE))))
    bolded_row_text = ' '.join([a.text for a in bolded_row_text]).strip()
    row_text = row.text.replace('\u200b', '').strip()
    if alltext:
        return len(bolded_row_text) > 0 and len(bolded_row_text) == len(row_text)
    else:
        return len(bolded_row_text) > 0


def is_italic(row):
    soup = BeautifulSoup(str(row), features='html.parser')
    italic = soup.find_all(lambda tag: tag.name == 'i')
    return len(italic) > 0


def is_centered(row):
    soup = BeautifulSoup(str(row), features='html.parser')
    return len(soup.find_all(lambda tag: tag.name != 'table' and (
            (tag.has_attr('align') and 'center' in str(tag['align'])) or tag.has_attr(
        'style') and 'text-align:center' in str(tag['style'])))) > 0


def find_left_margin(reg_row, td):
    # https://www.w3schools.com/cssref/pr_margin.asp
    pattern_matches = [
        re.findall('(margin-left|padding-left|text-indent):(-?\d+)', str(td)),
        re.findall(r'margin:-?\d+pt -?\d+pt -?\d+pt (-?\d+)pt', str(td)),
        re.findall(r'margin:-?\d+pt (-?\d+)pt -?\d+pt', str(td)),
        re.findall(r'margin:-?\d+pt (-?\d+)pt', str(td)),
        re.findall(r'margin:(-?\d+)pt', str(td))]
    for match in pattern_matches:
        c1 = sum([float(m[-1]) for m in match]) if len(match) > 0 else 0
        match = re.search(r'( *)\w', reg_row[0])
        c2 = match.group().count(' ') if match else 0
        if max(c1, c2) == 0:
            continue
        else:
            return max(c1, c2)
    return 0


# TODO MERGE FIND META TABLE INFO AND TABLE TITLE METHODS


def find_meta_table_info(table):
    table_multiplier, table_currency = '', ''
    multiplier_pattern = re.compile('(thousands|millions|billions|percentage)', re.IGNORECASE)
    currency_pattern = re.compile('([$€¥£])', re.IGNORECASE)
    current_element = table
    try:
        while current_element.previous_element is not None:

            if isinstance(current_element, NavigableString):
                current_element = current_element.previous_element
                continue

            # stop at first
            if re.search(multiplier_pattern, current_element.text) and len(table_multiplier) == 0:
                table_multiplier = re.search(multiplier_pattern, current_element.text).groups()[-1]

            if re.search(currency_pattern, current_element.text) and len(table_currency) == 0:
                table_currency = re.search(currency_pattern, current_element.text).groups()[-1]

            current_element = current_element.previous_element
    except:
        traceback.print_exc()

    return table_multiplier, table_currency


def find_table_title(table):
    priority_title, emergency_title = '', ''
    current_element = table
    try:

        # this is a check in case we moved to another table, it should exit
        while (table.text == current_element.text and current_element.name == 'table') \
                or current_element.name != 'table':

            # current element is the table, just move to previous element
            current_element = current_element.previous_element

            while (  # sometimes the previous element is a new line metacharacter (&nbsp or encoded as '\n') so skip
                    ((isinstance(current_element, NavigableString) and len(
                        current_element.replace('\n', '').strip()) == 0)
                     or ((not isinstance(current_element, NavigableString) and
                          # if previous element is a div, then it is enclosing what was in current element in previous iteration
                          # so just take the new element
                          ((len(current_element.contents) > 0

                            and (  # if the element is another tag, get text
                                    (not isinstance(current_element.contents[0], NavigableString) and
                                     len(current_element.contents[0].text.replace('\n', '').strip()) == 0)
                                    # if directly a string, no need to get text attribute first
                                    or (isinstance(current_element.contents[0], NavigableString) and
                                        len(current_element.contents[0].replace('\n', '').strip()) == 0)))
                           # sometimes it is another element that has no text (edge case)
                           or current_element.text == ''))))):
                current_element = current_element.previous_element

            # Sometimes the title will just be the Navigable String, so go to parent to capture the tag (which includes style):
            if isinstance(current_element, NavigableString):
                current_element = current_element.parent

            # TODO Current fix for the following (it takes the 't' only)
            '''<p style="font-family:'Times New Roman';font-size:10pt;margin:0pt;"><b style="font-weight:bold;">Consolidated Balance Shee</b><b style="font-weight:bold;">t</b></p>'''
            if len(current_element.text) == 1:
                while isinstance(current_element.previous_element, NavigableString):
                    current_element = current_element.previous_element.previous_element
                continue

            # TODO Also add is_colored function
            if is_bold(current_element) or is_centered(current_element) or is_italic(current_element):
                # sometimes the previous element is a detail of the title (i.e. (in thousands)), usually bracketed
                if re.search('^\((.*?)\)\*?$', current_element.text.strip()) \
                        or current_element.text.strip().replace('\u200b', '') == '' \
                        or re.search(fin_reg.date_regex, current_element.text.strip()) \
                        or (current_element.name == 'font' and re.search('^div$|^p$', current_element.parent.name)) \
                        or re.search('(Form 10-K|\d{2})', current_element.text.strip(), re.IGNORECASE):
                    continue
                else:
                    return unicodedata.normalize("NFKD", current_element.text).strip().replace('\n', '')

            elif re.search('The following table', current_element.text, re.IGNORECASE):
                emergency_title = current_element.text

        # if we reached here, then we haven't found bold/centered/italic
        if len(emergency_title) > 0:
            return emergency_title
        else:
            return 'No Table Title'
    except:
        traceback.print_exc()
        return 'No Table Title'


def flatten_dict(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = parent_key + sep + k if parent_key else k
        if isinstance(v, collections.abc.MutableMapping):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)


def scrape_pdf():
    file = 'https://reports.shell.com/annual-report/2019/servicepages/downloads/files/shell_annual_report_2019.pdf'
    tables = tabula.read_pdf(file, pages=196)
    pprint(table.to_string() for table in tables)


def scrape_txt_tables_from_url(url):
    pass


def scrape_xbrl_tables_from_url(url):
    pass


def scrape_html_tables_from_url(url, filing_date):
    response = requests.get(url)
    import data_scraping.unit_tests as tests
    table = tests.table_cash_flow_statement_2
    soup = BeautifulSoup(response.text, 'html.parser')
    all_in_one_dict = {'Yearly': {}, 'Quarterly': {}, '6 Months': {}, '9 Months': {}}

    month_abc_regex = r'Jan(?=uary)?|Feb(?=ruary)?|Mar(?=ch)?|Apr(?=il)?|May|Jun(?=e)?|Jul(?=y)?|Aug(?=ust)?|Sep(?=tember)?|Oct(?=ober)?|Nov(?=ember)?|Dec(?=ember)?'
    month_no_regex = r'1[0-2]|0?[1-9]'
    day_no_regex = r'3[01]|[12][0-9]|0?[1-9]'
    day_abc_regex = r'Mon(day)?|Tue(sday)?|Wed(nesday)?|Thur(sday)?|Fri(day)?|Sat(urday)?|Sun(day)?'
    year_regex_two = r'[0-9]{2}'
    year_regex_four = r'\d{4}'

    month_slash_day_slash_year_regex = r'((?:1[0-2]|0?[1-9])\/(?:3[01]|[12][0-9]|0?[1-9])\/(?:[0-9]{2})?[0-9]{2})'
    month_day_year_regex = r'({})\s+({}),?\s+({})'.format(month_abc_regex, day_no_regex, year_regex_two)
    flexible_month_day_year_regex = r'({}).*?({}).*?({})'.format(month_abc_regex, day_no_regex,
                                                                 year_regex_four)
    only_year_regex = r'^({})$'.format(year_regex_four)
    date_formats = r"(((1[0-2]|0?[1-9])\/(3[01]|[12][0-9]|0?[1-9])\/(?:[0-9]{2})?[0-9]{2})|((Jan(uary)?|Feb(ruary)?|Mar(ch)?|Apr(il)?|May|Jun(e)?|Jul(y)?|Aug(ust)?|Sep(tember)?|Oct(ober)?|Nov(ember)?|Dec(ember)?)\s+\d{1,2},\s+\d{4}))"

    if 'Inline XBRL Viewer' in soup.text:
        driver = webdriver.Chrome(ChromeDriverManager().install())
        driver.get(url)
        sleep(2)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')

    for table in soup.findAll('table'):
        columns = []
        dates = []
        header_found = False
        indented_list = []
        rows = table.find_all('tr')
        table_title, table_currency, table_multiplier = '', '', ''
        first_level = ''  # that's for whether the table is yearly of quarterly
        for index, row in enumerate(rows):

            reg_row = [ele.text for ele in row.find_all(lambda tag: tag.name == 'td' or tag.name == 'th')]
            reg_row = [unicodedata.normalize("NFKD", x).replace('\u200b', '') for x in reg_row]
            current_left_margin = find_left_margin(reg_row, row.findAll('td')[0])
            reg_row = [x.strip() for x in reg_row]
            reg_row[0] = reg_row[0].replace(':', '').replace('\n', ' ') if len(reg_row) > 0 else reg_row[0]

            # first, we want to construct the column header of our dable. We skip
            # - the empty rows i.e. those that have no table data 'td' (or those for which the td's are empty)
            # [length 0]
            # - the descriptive rows (i.e. 'dollars in millions', 'fiscal year', etc.) [length 1]
            # We only keep the useful headers i.e. those that include 'Month', 'Quarter', 'Year', a date...

            if not header_found:
                reg_row = list(filter(lambda x: x != "", reg_row))
                max_date_index = 0
                # if 'th' tag found or table data with bold, then found a potential header for the table
                if len(row.find_all('th')) != 0 or is_bold(row):
                    if len(columns) == 0:
                        columns = reg_row
                    else:
                        if len(columns) < len(reg_row):
                            ratio = int(len(reg_row) / len(columns))
                            col_len = len(columns)
                            for col_idx in range(col_len):
                                copy = []
                                for r in range(ratio):
                                    formatted_column = " ".join((columns[0] + ' ' + reg_row[0]).split())
                                    copy.append(formatted_column)
                                    reg_row.pop(0)
                                columns.extend(copy)
                                columns.pop(0)
                        else:
                            for r in reg_row:
                                for col in range(len(columns)):
                                    columns[col] = columns[col] + ' ' + r

                    # sometimes same title (i.e. if table is split in two pages, they repeat page title twice
                    table_multiplier, table_currency = find_meta_table_info(table)
                    table_title = find_table_title(table=table)
                    # TODO if table title includes 'parent company', skip
                    if table_multiplier == 'percentage':  # TODO ?
                        break
                        # normalizing to units in 1000s
                    elif re.search('million', table_multiplier, re.IGNORECASE):
                        table_multiplier = 1000
                    elif re.search('billion', table_multiplier, re.IGNORECASE):
                        table_multiplier = 1000000
                    elif re.search('thousand', table_multiplier, re.IGNORECASE):
                        table_multiplier = 1
                    else:
                        table_multiplier = 0.001

                    format_match = False
                    dates = []
                    only_year = False
                    for col in columns:
                        # Filing for June 30, 2019:
                        # Fiscal year 2019  September   December    March   June --> september and december are 2018
                        if only_year:
                            break
                        match = re.search(flexible_month_day_year_regex, col) or re.search(
                            month_slash_day_slash_year_regex, col) or re.search(only_year_regex, col)
                        if match:
                            for dts in ['%b %d %Y', '%m/%d/%y', '%Y']:
                                try:
                                    if re.search(only_year_regex,
                                                 col):  # when only year, need to make exception because we'll only keep the data from that year (we won't do the other since we don't have filing dates yet)
                                        dates.append(filing_date)
                                        format_match = True
                                        only_year = True
                                        break
                                    # print(match.groups())
                                    col_formatted_date = datetime.strptime(' '.join(match.groups()), dts).date()
                                    col_formatted_date = datetime(col_formatted_date.year, col_formatted_date.month,
                                                                  col_formatted_date.day)
                                    # ascending = True if dates[0] < dates[1] else False
                                    if col_formatted_date > filing_date:
                                        col_formatted_date = datetime(col_formatted_date.year - 1,
                                                                      col_formatted_date.month,
                                                                      col_formatted_date.day)
                                    dates.append(col_formatted_date)
                                    format_match = True
                                    break
                                except:
                                    continue

                    # relevant_indexes = [i for i, x in enumerate(columns) if not re.search(r'Six|Nine|Change', x, re.IGNORECASE)]
                    # if only_year:
                    #     try:
                    #         relevant_indexes = columns.index(str(filing_date.year))
                    #     except:
                    #         print('NOT IN LISTTTTTT {}'.format(table_title))
                    # print(relevant_indexes)
                    if len(columns) > 0 and format_match:
                        header_found = True
                        header_index = index
                        # first_bolded_index = next(i for i, v in enumerate(indented_list) if v[2])


                else:
                    continue

            elif header_found and len(reg_row) > 0:

                indices = [i for i, x in enumerate(reg_row) if re.search(r'\d', x)]
                reg_row = [reg_row[0]] + [re.sub(r'(^-$)|(^—$)', '0', x) for x in reg_row[1:]]

                reg_row = [reg_row[0]] + [re.sub(r'\(', '-', x) for x in reg_row[1:]]
                reg_row = [reg_row[0]] + [re.sub("[^0-9-]", "", r) for r in reg_row[1:]]
                reg_row = [reg_row[0]] + [re.sub("^-$", "", r) for r in reg_row[1:]]  # for NOT the minus sign
                reg_row = [reg_row[0]] + [re.sub(r',', '', x) for x in reg_row[1:]]
                reg_row = list(filter(lambda x: x != "", reg_row))

                if len(''.join(reg_row).strip()) == 0:
                    continue

                while len(indented_list) > 0 and current_left_margin <= indented_list[-1][0]:

                    if current_left_margin == indented_list[-1][0]:

                        if indented_list[-1][2] or indented_list[-1][1].isupper():  # if last element of list is bold
                            if is_bold(row, alltext=True) or row.text.isupper():  # if current row is bold
                                indented_list.pop()  # remove that last element of list (new bold overrides old bold)
                                break  # and stop popping
                            else:
                                break  # otherwise, just subentry so don't pop
                    indented_list.pop()  # pop (because the most recent category is the element itself so we want to replace it)

                try:
                    # if re.search(r'Total', reg_row[0], re.IGNORECASE):
                    #     indented_list.pop()
                    for i in range(len(indented_list)):
                        if re.search(r'^{}$'.format(reg_row[0].split('Total ')[-1]),  # current entry
                                     indented_list[-i][1],  # last category
                                     re.IGNORECASE):
                            indented_list.pop()
                except Exception:
                    traceback.print_exc()

                indented_list.append((current_left_margin, reg_row[0], is_bold(row, alltext=True)))
                current_category = '_'.join([x[1] for x in indented_list])

                if len(reg_row) > 1:  # not category column:
                    try:
                        for index, col in enumerate(columns):

                            if re.search(r'Three|Quarter', table_title + col, re.IGNORECASE):
                                first_level = 'Quarterly'
                            elif re.search(r'Six', table_title + col, re.IGNORECASE):
                                first_level = '6 Months'
                            elif re.search(r'Nine', table_title + col, re.IGNORECASE):
                                first_level = '9 Months'
                            elif re.search(r'Change', table_title + col, re.IGNORECASE):
                                continue
                            else:
                                first_level = 'Yearly'

                            for date in dates:
                                if date not in all_in_one_dict[first_level].keys():
                                    all_in_one_dict[first_level][date] = {}
                                if table_title not in all_in_one_dict[first_level][date].keys():
                                    all_in_one_dict[first_level][date][table_title] = {}
                            if only_year:
                                # index = columns.index(str(filing_date.year))
                                index = dates.index(filing_date)
                            if current_category not in all_in_one_dict[first_level][dates[index]][table_title].keys():
                                all_in_one_dict[first_level][dates[index]][table_title][current_category] = float(
                                    re.sub('^$', '0', reg_row[index + 1]))  # * table_multiplier
                                # pprint(all_in_one_dict[first_level][dates[index]][table_title])
                            if only_year:
                                break
                    except:
                        # print('EXCEPTION INDEX! for title {} and row {}, with col {}'.format(table_title, reg_row, col))
                        pass

    return all_in_one_dict


def scrape_tables_from_url(url, filing_date):
    extension = url.split('.')[-1]

    if extension == 'xml':
        return scrape_xbrl_tables_from_url(url)
    elif extension == 'htm':
        return scrape_html_tables_from_url(url, filing_date)
    elif extension == 'txt':
        return scrape_txt_tables_from_url(url)


def normalization_iteration(iteration_count, input_dict, master_dict, visited_data_names, year,
                            flexible_sheet=False, flexible_entry=False):
    for title, table in input_dict.items():

        for scraped_name, scraped_value in flatten_dict(table).items():
            found_and_done = False
            # for visited_data in visited_data_names[year]:
            for normalized_category, pattern_string in flatten_dict(fin_reg.financial_entries_regex_dict).items():
                if found_and_done:
                    break
                # if you're a flexible sheet, the sheet we're checking at least shouldn't match the other concerning statements (i.e. depreciation and amortization's pattern in balance sheet regex shouldn't match cash flow statement change in depreciation and amortization)
                # TODO for now, we're only allowing balance sheet and income statement together (because of shares outstanding)
                if (flexible_sheet and (('Balance Sheet' in normalized_category.split('_')[0]
                                         and not re.search(
                            r'{}'.format(fin_reg.cash_flow_statement_regex), title,
                            re.IGNORECASE))
                                        or ('Income Statement' in normalized_category.split('_')[0]
                                            and not re.search(
                                    r'{}'.format(fin_reg.cash_flow_statement_regex),
                                    title, re.IGNORECASE))
                                        or ('Cash Flow Statement' in normalized_category.split('_')[0]
                                            and not re.search(
                                    r'{}|{}'.format(fin_reg.balance_sheet_regex, fin_reg.income_statement_regex), title,
                                    re.IGNORECASE)))

                        # if you're not a flexible sheet, the sheet we're checking must match regex sheet
                        or ((not flexible_sheet)
                            and (('Balance Sheet' in normalized_category.split('_')[0] and re.search(
                                    fin_reg.balance_sheet_regex, title, re.IGNORECASE))
                                 or ('Income Statement' in normalized_category.split('_')[0] and re.search(
                                            fin_reg.income_statement_regex, title, re.IGNORECASE))
                                 or ('Cash Flow Statement' in normalized_category.split('_')[0] and re.search(
                                            fin_reg.cash_flow_statement_regex, title, re.IGNORECASE))))):

                    # an entry is not flexible if it should match a hardcoded pattern
                    pattern_string = '^' + pattern_string
                    if ((not flexible_entry) and re.search(
                            pattern_string, title + '_' + scraped_name, re.IGNORECASE)):
                        # print('Found pattern match {} for scraped name {}'.format(pattern_string, scraped_name))

                        data_name = {'Iteration Count': str(iteration_count),
                                     'Hardcoded': True,
                                     'Table Title': title,
                                     'Scraped Name': scraped_name,
                                     'Whole Name': normalized_category,
                                     'Pattern String': pattern_string}

                        if math.isnan(master_dict[normalized_category]):
                            if year not in visited_data_names.keys():
                                visited_data_names[year] = []
                            pattern_matched = False
                            for el in visited_data_names[year]:  # if I already found it for this year, then skip

                                if re.search(el['Pattern String'], scraped_name, re.IGNORECASE):
                                    # TODO: check following bug fix: you should find it this year, but the pattern should match same table title
                                    if re.search(el['Table Title'], title, re.IGNORECASE):
                                        pattern_matched = True
                                    break
                            if not pattern_matched:
                                master_dict[normalized_category] = scraped_value
                                visited_data_names[year].append(data_name)  # that table takes ownership for the data
                            break

                    # otherwise it is flexible if it should just match the category
                    # (i.e. current assets, operating expenses...)
                    # TODO since the outer loop is over each name, and we just need to go once for each category,
                    # we need to do something about it (inefficient)

                    if flexible_entry:
                        if found_and_done:
                            break
                        # make normalized category into a regex
                        categories_balance_sheet = [
                            # Balance Sheet
                            re.compile(r'((.*?_?)Current Assets)', re.IGNORECASE),
                            re.compile('((.*?_?)Current Liabilities)', re.IGNORECASE),
                            re.compile(r'((.*?_?)Non-? ?Current Assets)', re.IGNORECASE),
                            re.compile('((.*?_?)Non-? ?Current Liabilities)', re.IGNORECASE),
                            re.compile('((.*?)(?!.*Liabilities)Shareholders[’\'] Equity)', re.IGNORECASE),
                            re.compile(r'((.*?_?)Assets)', re.IGNORECASE),
                            re.compile('((.*?)Liabilities(?!.*Shareholders[’\'] Equity))', re.IGNORECASE),
                            # Income Statement
                            re.compile('((.*?)Revenues)', re.IGNORECASE),
                            re.compile('((.*?)Operating Expenses)', re.IGNORECASE),
                            # re.compile('((.*?)Other Expenses)', re.IGNORECASE),
                            # Cash Flow Statement
                            re.compile('((.*?)Operating Activities(?!.*Net.*_))', re.IGNORECASE),
                            re.compile('((.*?)Investing Activities(?!.*Net.*_))', re.IGNORECASE),
                            re.compile('((.*?)Financing Activities(?!.*Net.*_))', re.IGNORECASE)
                        ]

                        for cat in categories_balance_sheet:
                            if found_and_done:
                                break
                            match = re.search(cat, normalized_category)
                            if match and re.search(cat, scraped_name):
                                # print(match.groups())
                                category = match.groups()[0]
                                already_have_it = False
                                for el in visited_data_names[year]:
                                    # make scraped name into a regex afterwards
                                    if re.search(el['Pattern String'], scraped_name, re.IGNORECASE):
                                        same_sheet = False
                                        for regex_title in [fin_reg.balance_sheet_regex, fin_reg.income_statement_regex,
                                                            fin_reg.cash_flow_statement_regex]:
                                            if re.search(regex_title, title, re.IGNORECASE) \
                                                    and re.search(regex_title, el['Table Title'], re.IGNORECASE):
                                                same_sheet = True

                                        if same_sheet:
                                            already_have_it = True

                                if not already_have_it:
                                    # TODO BIG CHANGE HERE
                                    name = titlecase(scraped_name.split('_')[-1])
                                    master_dict[category + '_' + name] = scraped_value
                                    name_regex = r''
                                    for word in scraped_name.split('_'):
                                        name_regex += r'(?=.*{}(?!.*[_]))'.format(word)
                                    visited_data_names[year].append({'Iteration Count': str(iteration_count),
                                                                     'Hardcoded': False,
                                                                     'Table Title': title,
                                                                     'Scraped Name': name,
                                                                     'Whole Name': category + '_' + name,
                                                                     'Pattern String': name_regex})

                                found_and_done = True
                                break
    return visited_data_names, master_dict


def normalize_tables(input_dict, visited_data_names, year):
    # pprint(input_dict)
    master_dict = {}

    for normalized_category, pattern_string in flatten_dict(fin_reg.financial_entries_regex_dict).items():
        master_dict[normalized_category] = np.nan

    visited_data_names, master_dict = normalization_iteration(0, input_dict, master_dict,
                                                              visited_data_names, year, flexible_sheet=False,
                                                              flexible_entry=False)

    # then we want to give priority to the elements that strictly match our regex patterns,
    # but not in the consolidated financial statements
    visited_data_names, master_dict = normalization_iteration(1, input_dict, master_dict,
                                                              visited_data_names, year, flexible_sheet=True,
                                                              flexible_entry=False)

    # finally, we want to give priority to the rest of the elements (i.e. those that do not
    # match our regex patterns) that are in the consolidated financial statements
    # TODO bug: duplicate entries despite pattern match check
    visited_data_names, master_dict = normalization_iteration(2, input_dict, master_dict,
                                                              visited_data_names, year, flexible_sheet=False,
                                                              flexible_entry=True)
    pprint(master_dict)
    # TODO Final Standardization, Fill in Differences!
    # If Interest Income/Expense in revenues, then readjust Net Sales (add back)
    # total noncurrent = total - total current  for assets, liabilities
    # if total noncurrent nan, else total = total current + total noncurrent

    # master_dict['Balance Sheet_Assets_Non Current Assets_Total Non Current Assets'] = master_dict[
    #                                                                                       'Balance Sheet_Assets_Total Assets'] - \
    #                                                                                   master_dict[
    #                                                                                       'Balance Sheet_Assets_Current Assets_Total Current Assets']
    #
    # master_dict[
    #     'Balance Sheet_Liabilities and Shareholders\' Equity_Liabilities_Non Current Liabilities_Total Non Current Liabilities'] = \
    #     master_dict['Balance Sheet_Liabilities and Shareholders\' Equity_Liabilities_Total Liabilities'] - master_dict[
    #         'Balance Sheet_Liabilities and Shareholders\' Equity_Liabilities_Current Liabilities_Total Current Liabilities']

    return visited_data_names, flatten_dict(unflatten(master_dict))


def unflatten(dictionary):
    resultDict = dict()
    for key, value in dictionary.items():
        parts = key.split("_")
        d = resultDict
        for part in parts[:-1]:
            if part not in d:
                d[part] = dict()
            d = d[part]
        d[parts[-1]] = value
    return resultDict


# make sure delta between bigger and smaller is 3 months i.e. 3 and 6, 6 and 9, 9 and 12
def quarterlize_statements(smaller_period_df, bigger_period_df, existing_quarters):
    df_output = pd.DataFrame()
    for bigger_period_index, bigger_period_date in enumerate(bigger_period_df.columns):
        if bigger_period_date not in existing_quarters:
            try:
                closest_smaller_period_index = next(
                    index for (index, item) in enumerate(smaller_period_df.columns) if item < bigger_period_date)
            except:
                continue
            # print(closest_smaller_period_index)
            df_output[bigger_period_date] = bigger_period_df[bigger_period_date] - smaller_period_df.iloc[:,
                                                                                   closest_smaller_period_index]
    # print(smaller_period_df.to_string())
    return df_output


def scrape_financial_statements(ticker, how_many_years=2, how_many_quarters=8):
    path = '{}/{}.xlsx'.format(config.FINANCIAL_STATEMENTS_DIR_PATH, ticker)
    log_folder_path = os.path.join(config.DATA_DIR_PATH, 'logs')
    if not os.path.exists(log_folder_path):
        os.mkdir(log_folder_path)
    company_log_path = os.path.join(log_folder_path, ticker)
    if not os.path.exists(company_log_path):
        os.mkdir(company_log_path)

    dictio = {}
    cik = get_company_cik(ticker)
    doc_links_yearly = get_filings_urls_first_layer(cik, '10-K')
    doc_links_quarterly = get_filings_urls_first_layer(cik, '10-Q')

    filings_dictio_yearly = get_filings_urls_second_layer(doc_links_yearly)

    year_dates = set([tuple_1
                      for key, value in filings_dictio_yearly.items()
                      for tuple_1, tuple_2 in value])

    filings_dictio_quarterly = get_filings_urls_second_layer(doc_links_quarterly)

    filing_dictio = {}

    for i in range(2):
        for title, links_list in (filings_dictio_yearly.items() if i == 0 else filings_dictio_quarterly.items()):
            if title not in filing_dictio.keys():
                filing_dictio[title] = []
            filing_dictio[title].extend(links_list)
        missing_dates_links = []
        existing_dates = excel.read_dates_from_csv(path,
                                                   config.balance_sheet_yearly if i == 0 else config.balance_sheet_quarterly)
        for x, y in (filings_dictio_yearly['HTML Links'] if i == 0 else filings_dictio_quarterly['HTML Links']):
            formatted_date = datetime.strptime(x, '%Y-%m-%d')
            if formatted_date not in existing_dates and formatted_date not in [x for x, y in missing_dates_links]:
                missing_dates_links.append((formatted_date, y))
        missing_dates_links.sort(key=lambda tup: tup[0], reverse=True)

        # pprint(missing_dates_links)
        for index, (filing_date, link) in enumerate(missing_dates_links):
            try:
                if (index > how_many_years - 1 and i == 0) or (index > how_many_quarters - 1 and i == 1):
                    break
                print(filing_date, link)
                output = scrape_tables_from_url(link, filing_date)
                pprint(output)
                for sheet_period, sheet_dict in output.items():
                    if sheet_period not in dictio.keys():
                        dictio[sheet_period] = {}
                    for year, title_dict in sheet_dict.items():
                        # if we don't have the year in our dictio that collects everything, just add all and go to next year of the output
                        if year not in dictio[sheet_period].keys():
                            dictio[sheet_period][year] = title_dict
                            continue

                        #  else, we have a year, so we add up those two dicts together
                        for title, last_layer in title_dict.items():  # title / key:float
                            if title not in dictio[sheet_period][year].keys():
                                dictio[sheet_period][year][title] = last_layer

                            else:
                                dictio[sheet_period][year][title].update(last_layer)

            except Exception:
                traceback.print_exc()
        # pprint(dictio)

    log = open(os.path.join(company_log_path, 'scraped_dictio.txt'), "w")
    # try:
    #     # print(dictio, file=log)
    # except:
    #     pass
    financials_dictio = {}

    for sheet_period, sheet_dict in dictio.items():
        visited_data_names = {}
        if sheet_period not in financials_dictio.keys():
            financials_dictio[sheet_period] = {}
        for year, title_dict in sheet_dict.items():
            if year not in financials_dictio[sheet_period].keys():
                financials_dictio[sheet_period][year] = {}
            visited_data_names, financials_dictio[sheet_period][year] = normalize_tables(title_dict, visited_data_names,

                                                                                         year)
        log = open(os.path.join(company_log_path, '{}_normalized_dictio.txt'.format(sheet_period)), "w")
        print(visited_data_names, file=log)

    sheet_names = [config.balance_sheet_name, config.income_statement_name, config.cash_flow_statement_name]
    dfs_dictio = {}
    for sheet_name in sheet_names:
        if sheet_name not in dfs_dictio.keys():
            dfs_dictio[sheet_name] = {}
        first_balance_sheet_df = pd.DataFrame()
        visited_first_balance_sheet = False
        for sheet_period, sheet_dict in financials_dictio.items():
            if sheet_period not in dfs_dictio.keys():
                dfs_dictio[sheet_name][sheet_period] = pd.DataFrame()
                #  collections.OrderedDict(
            diction = collections.OrderedDict({i.strftime('%Y-%m-%d'): collections.OrderedDict({
                (j.split('_')[1], j.split('_')[-1])
                if 'Balance Sheet' not in sheet_name else (
                    (j.split('_')[1],
                     j.split('_')[2] if (len(j.split('_')) > 2) else ' ',
                     j.split('_')[-1]))
                : sheet_dict[i][j]
                for j in sheet_dict[i].keys() if j.split('_')[0] in sheet_name  # sheet name
            }) for i in sheet_dict.keys()})  # )  # date

            if sheet_name == config.balance_sheet_name and sheet_period in [config.yearly, config.quarterly] \
                    and not visited_first_balance_sheet:
                first_balance_sheet_df = pd.DataFrame.from_dict(diction)
                # print(first_balance_sheet_df.to_string())
                visited_first_balance_sheet = True
            else:
                df = pd.DataFrame.from_dict(diction)
                # pprint(diction)
                if sheet_name == config.balance_sheet_name and sheet_period in [config.yearly, config.quarterly]:
                    if visited_first_balance_sheet:
                        # print(df.to_string())
                        # df = pd.concat([first_balance_sheet_df, df], axis=0)
                        # first_balance_sheet_df.fillna(df, inplace=True)
                        # df.fillna(first_balance_sheet_df, inplace=True)
                        df = pd.concat([df, first_balance_sheet_df], axis=1, join='outer')
                        # print(df.to_string())
                        df = df.groupby(lambda x: x, axis=1).max()
                        df_yearly = df[[x for x in df.columns if x in year_dates]]
                        for datafrme, period in zip([df_yearly, df], [config.yearly, config.quarterly]):
                            datafrme = datafrme.reindex(sorted(datafrme.columns, reverse=True), axis=1)
                            datafrme.dropna(axis=0, how='all', inplace=True)
                            datafrme = datafrme.loc[:, datafrme.any()]
                            if not datafrme.empty:
                                # dfs_dictio[sheet_name][sheet_period] = datafrme
                                excel.save_into_csv(path, datafrme, '{} ({})'.format(sheet_name, period))
                    break

                df = df.reindex(sorted(df.columns, reverse=True), axis=1)
                df.dropna(axis=0, how='all', inplace=True)
                df = df.loc[:, df.any()]
                if not df.empty:
                    # dfs_dictio[sheet_name][sheet_period] = df
                    excel.save_into_csv(path, df, '{} ({})'.format(sheet_name, sheet_period))

    #  this is to standardize cumulated 3 6 9 12 months (only for cash flow statement for now)
    for quarterly_statement in [config.cash_flow_statement_quarterly, config.income_statement_quarterly]:

        try:
            quarterly_df = pd.read_excel(path, quarterly_statement, index_col=[0, 1])
            # quarterly_df = dfs_dictio[]
        except:
            continue
        temp_statements = config.income_statements if quarterly_statement == config.income_statement_quarterly else config.cash_flow_statements
        for i, item in enumerate(temp_statements):
            try:
                smaller_period_df = pd.read_excel(path, item, index_col=[0, 1])
                bigger_period_df = pd.read_excel(path, temp_statements[i + 1], index_col=[0, 1])
            except:
                continue

            quarterly_df = pd.concat([quarterly_df,
                                      quarterlize_statements(smaller_period_df, bigger_period_df,
                                                             quarterly_df.columns)],
                                     axis=1)
        quarterly_df = quarterly_df.reindex(sorted(quarterly_df.columns, reverse=True), axis=1)
        print(quarterly_df.to_string())
        excel.save_into_csv(path, quarterly_df, quarterly_statement, overwrite_sheet=True)

    #  then you can remove the 6 and 9 months sheets
    book = load_workbook(path)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
        for sheet_name in writer.book.sheetnames:
            if config.six_months in sheet_name or config.nine_months in sheet_name:
                writer.book.remove(writer.book[sheet_name])
        writer.book.save(path)
        writer.close()

    #  finally, paint alternate rows to excel file
    book = load_workbook(path)
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        if os.path.exists(path):
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    for sheet_name in writer.book.sheetnames:
        sheet = writer.book[sheet_name]
        df = pd.read_excel(path, sheet_name)

        start_column_space_for_index = 3 if config.balance_sheet_name in sheet_name else 2
        end_column_space_for_index = len(df.columns) + start_column_space_for_index - 1 \
            if config.balance_sheet_name not in sheet_name else len(df.columns) + start_column_space_for_index - 2
        row_space_for_index = 2
        for i in range(start_column_space_for_index, end_column_space_for_index):  # x-axis
            for j in range(row_space_for_index, len(df) + row_space_for_index):  # y-axis
                color = 'EEEEEE' if j % 2 == 0 else 'FFFFFF'  # integer odd-even alternation
                sheet.cell(row=j, column=i).fill = PatternFill(start_color=color,
                                                               fill_type='solid')

    writer.book.save(path)
    writer.close()


def save_stock_prices(stock, start=datetime(1970, 1, 1), end=datetime.now()):
    df = web.DataReader(stock.replace('.', '-'), data_source='yahoo', start=start, end=end)
    df['Pct Change'] = df['Adj Close'].pct_change()
    df.index = df.index + timedelta(days=1) - timedelta(seconds=1)
    path = os.path.join(config.STOCK_PRICES_DIR_PATH, '{}.xlsx'.format(stock))
    excel.save_into_csv(path, df, overwrite_sheet=True)
    return df


def get_technical_indicators(stock):
    df = excel.read_df_from_csv(stock, config.technical_indicators_name)
    if df.empty:
        df = save_stock_prices(stock)
        df = ta.add_all_ta_features(
            df, open="Open", high="High", low="Low", close="Adj Close", volume="Volume", fillna=True)
        path = '{}/{}.xlsx'.format(config.FINANCIAL_STATEMENTS_DIR_PATH, stock)
        excel.save_into_csv(path, df, config.technical_indicators_name)
        return df
    else:
        return df


def save_gnp_price_index():
    url = 'https://fred.stlouisfed.org/graph/fredgraph.csv?bgcolor=%23e1e9f0&chart_type=line&drp=0&fo=open%20sans&' \
          'graph_bgcolor=%23ffffff&height=450&mode=fred&recession_bars=on&txtcolor=%23444444&ts=12&tts=12&width=1168&nt' \
          '=0&thu=0&trc=0&show_legend=yes&show_axis_titles=yes&show_tooltip=yes&id=A001RG3A086NBEA&scale=left&cosd=1929' \
          '-01-01&coed=2019-01-01&line_color=%234572a7&link_values=false&line_style=solid&mark_type=none&mw=3&lw=2&ost=' \
          '-99999&oet=99999&mma=0&fml=a&fq=Annual&fam=avg&fgst=lin&fgsnd=2009-06-01&line_index=1&transformation=lin&vin' \
          'tage_date=2020-07-10&revision_date=2020-07-10&nd=1929-01-01'

    urllib.request.urlretrieve(url, 'temp.csv')
    df = pd.read_csv('temp.csv')
    df.columns = ['Date', 'GNP Price Index']
    df.set_index(df.columns[0], inplace=True)
    df.index = pd.to_datetime(df.index, format='%Y-%m-%d')
    # df.index = df.index - timedelta(days=1)

    with pd.option_context('display.max_rows', None, 'display.max_columns', None):  # more options can be specified also
        print(df)

    existing_df = excel.read_df_from_csv(config.MACRO_DATA_FILE_PATH, config.yearly_factors)
    final_df = existing_df.merge(df, how='outer', left_index=True, right_index=True)
    excel.save_into_csv(config.MACRO_DATA_FILE_PATH, final_df, config.yearly_factors)
    os.remove('temp.csv')


def save_factors_data(url, factor, period, skiprows):
    if len(url) == 0:  # that's for CAPM
        ff3 = pd.read_excel('{}/{}.xlsx'.format(config.FACTORS_DIR_PATH, 'Fama-French 3 Factor'),
                            sheet_name=period, index_col=0)
        capm = ff3[['Mkt-RF', 'RF']]
        excel.save_into_csv('{}/{}.xlsx'.format(config.FACTORS_DIR_PATH, factor), capm, sheet_name=period)
        return

    urllib.request.urlretrieve(url, 'fama_french.zip')
    zip_file = zipfile.ZipFile('fama_french.zip', 'r')
    zip_file.extractall()
    zip_file.close()

    file_name = next(file for file in os.listdir('.') if re.search('F-F', file))
    ff_factors = pd.read_csv(file_name, skiprows=skiprows, index_col=0)
    # We want to find out the row with NULL value. We will skip these rows
    ff_row = np.array(ff_factors.isnull().any(1)).nonzero()[0].tolist()
    ff_row = np.insert(ff_row, 0, 0, axis=0)

    # TODO waw hahaa worst hardcode
    temp_ff_row = ff_row
    for idx, val in enumerate(temp_ff_row):
        if abs(val - ff_row[idx - 1]) <= 1:
            ff_row = np.delete(ff_row, idx - 1)
            break

    first_pass = True
    formats = {'Daily': {'Format': '%Y%m%d',
                         'Offset': timedelta(days=1) - timedelta(seconds=1)},
               'Weekly': {'Format': '%Y%m%d',
                          'Offset': timedelta(days=1) - timedelta(seconds=1)},
               'Monthly': {'Format': '%Y%m',
                           'Offset': pd.offsets.MonthEnd()},
               'Yearly': {'Format': '%Y',
                          'Offset': pd.offsets.YearEnd()}}
    for idx, nrow in enumerate(ff_row[1:]):

        # Read the csv file again with skipped rows
        skiprows = skiprows + 3 if not first_pass else skiprows
        ff_factors = pd.read_csv(file_name,
                                 skiprows=ff_row[idx] + skiprows,
                                 nrows=nrow - ff_row[idx],
                                 index_col=0)

        ff_factors.index = [str(x).strip() for x in ff_factors.index.to_list()]
        cur_period = period.split('/')[0] if first_pass else period.split('/')[-1]
        format = formats[cur_period]['Format']

        try:
            ff_factors.index = pd.to_datetime(ff_factors.index, format=format)
        except:  # TODO hard fix, because of the 'copyright' row at bottom, filter later
            ff_factors = ff_factors[:-1]
            ff_factors.index = pd.to_datetime(ff_factors.index, format=format)

        if re.search('Monthly|Yearly', period, re.IGNORECASE):
            ff_factors.index = ff_factors.index + timedelta(days=1) - timedelta(seconds=1)

        ff_factors.index = ff_factors.index + formats[cur_period]['Offset']
        ff_factors = ff_factors.apply(lambda x: x / 100)
        path = os.path.join(config.FACTORS_DIR_PATH, '{}.xlsx'.format(factor))

        for col in ff_factors.columns:
            ff_factors.rename(columns={col: col.strip()}, inplace=True)

        if 'Carhart' in factor:
            ff3 = pd.read_excel('{}/{}.xlsx'.format(config.FACTORS_DIR_PATH, 'Fama-French 3 Factor'),
                                sheet_name=cur_period, index_col=0)
            ff_factors.rename(columns={'Mom': 'MOM'}, inplace=True)
            ff_factors = pd.merge(ff3, ff_factors, left_index=True, right_index=True)

        excel.save_into_csv(filename=path, df=ff_factors, sheet_name=cur_period, overwrite_sheet=True)
        first_pass = False

    os.remove(file_name)
    os.remove('fama_french.zip')

    return ff_factors


three_factors_url_daily = 'https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_Factors_daily_CSV.zip'
three_factors_url_weekly = 'https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_Factors_weekly_CSV.zip'
three_factors_url_monthly_yearly = "https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_Factors_CSV.zip"
five_factors_url_daily = 'https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_5_Factors_2x3_daily_CSV.zip'
five_factors_url_weekly = ''
five_factors_url_monthly_yearly = 'https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Research_Data_5_Factors_2x3_CSV.zip'
momentum_factor_url_daily = 'https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Momentum_Factor_daily_CSV.zip'
momentum_factor_url_monthly_yearly = 'https://mba.tuck.dartmouth.edu/pages/faculty/ken.french/ftp/F-F_Momentum_Factor_CSV.zip'
carhart_four_factor_url_monthly = 'https://breakingdownfinance.com/wp-content/uploads/2019/07/Carhart-4-factor-data.xlsx'

factors_inputs = [
    (three_factors_url_daily, 'Fama-French 3 Factor', 'Daily', 3),
    (three_factors_url_weekly, 'Fama-French 3 Factor', 'Weekly', 3),
    (three_factors_url_monthly_yearly, 'Fama-French 3 Factor', 'Monthly/Yearly', 3),
    (five_factors_url_daily, 'Fama-French 5 Factor', 'Daily', 3),
    (five_factors_url_monthly_yearly, 'Fama-French 5 Factor', 'Monthly/Yearly', 3),
    (momentum_factor_url_daily, 'Carhart 4 Factor', 'Daily', 13),
    (momentum_factor_url_monthly_yearly, 'Carhart 4 Factor', 'Monthly/Yearly', 13),
    ('', 'CAPM', 'Daily', 0), ('', 'CAPM', 'Weekly', 0),
    ('', 'CAPM', 'Monthly', 0), ('', 'CAPM', 'Yearly', 0)
]


def testing():
    pprint(scrape_html_tables_from_url(
        'https://www.sec.gov/ix?doc=/Archives/edgar/data/1018724/000101872420000010/amzn-20200331x10q.htm',
        datetime(2019, 12, 31)))


if __name__ == '__main__':

    if not os.path.exists(config.DATA_DIR_PATH):
        os.mkdir(config.DATA_DIR_PATH)
    if not os.path.exists(config.MARKET_TICKERS_DIR_PATH):
        os.mkdir(config.MARKET_TICKERS_DIR_PATH)
    if not os.path.exists(config.FINANCIAL_STATEMENTS_DIR_PATH):
        os.mkdir(config.FINANCIAL_STATEMENTS_DIR_PATH)
    if not os.path.exists(config.FACTORS_DIR_PATH):
        os.mkdir(config.FACTORS_DIR_PATH)
    if not os.path.exists(os.path.join(config.DATA_DIR_PATH, 'stock_prices')):
        os.mkdir(os.path.join(config.DATA_DIR_PATH, 'stock_prices'))

    # for url, file, sheet, skiprow in factors_inputs:
    #     save_factors_data(url, file, sheet, skiprow)
    with open("{}/djia30_tickers.pickle".format(config.MARKET_TICKERS_DIR_PATH), "rb") as f:
        tickers = pickle.load(f)
        for ticker in tickers[1:5]:
            try:
                scrape_financial_statements(ticker, how_many_years=1, how_many_quarters=0)
            except:
                continue
        # save_stock_prices(ticker)
        # get_technical_indicators(ticker)
        # scrape_financial_statements(ticker, how_many_years=5, how_many_quarters=0)
    # scrape_financial_statements(ticker, '10-Q')

    # ff_factors = get_beta_factors()
    # print(ff_factors)
    # testing()
    # scrape_pdf()
    # save_factors_data(momentum_factor_url_daily)
